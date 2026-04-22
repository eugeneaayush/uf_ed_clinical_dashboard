"""
UF Adult ED Split-Flow Back Test
--------------------------------
Simulates how actual Adult ED encounters (7/1/2025 - 4/19/2026) would have
performed under the proposed Split-Flow operating model. Writes a multi-sheet
Excel report with baseline, simulated, delta, and by-lane views.
"""
import sys
sys.stdout.reconfigure(encoding="utf-8")

import numpy as np
import pandas as pd
from pathlib import Path

ROOT = Path(r"C:\Users\atul0001\split-flow-backtest")
SRC  = ROOT / "Split Flow Back Test.xlsx"
OUT  = ROOT / "Split Flow Back Test - Results.xlsx"

# ---------- Parameters (Split-Flow assumptions) ----------
WINDOW_START = 8          # 08:00
WINDOW_END   = 1          # 01:00 (exclusive upper; hours 8..23 + 0)
IN_WINDOW_HOURS = list(range(WINDOW_START, 24)) + [0]

DOOR_TO_MD_BASE_CAP = 15            # normal cap
LWBS_CAPTURE_SUPERTRACK = 0.75      # 75% of ST-eligible LWBS retained (REVISED)
LWBS_CAPTURE_MAIN_ESI2  = 0.60      # 60% of ESI-2 LWBS retained (REVISED)
LWBS_CAPTURE_RED        = 1.00      # ESI-1 always bedded
MD_TO_DISP_FLAT_REDUX_ST = 45       # 45 min flat reduction for Super Track
MD_TO_DISP_FLAT_REDUX_MAIN = 20     # 20 min flat reduction for Main ED
MIN_MD_TO_DISP = 30                 # Floor for workup time

# ---------- Load and dedupe ----------
print("Loading data...")
df = pd.read_excel(SRC, sheet_name="Query 1")
enc = df.drop_duplicates(subset="Encounter # (CSN)").copy()
adult = enc[enc["ED Location"] == "ADULT ED"].copy().reset_index(drop=True)
print(f"  Distinct ADULT ED encounters: {len(adult):,}")

adult["Arrival DateTime"] = pd.to_datetime(adult["Arrival DateTime"])
adult["ArrHour"] = adult["Arrival DateTime"].dt.hour
adult["InWindow"] = adult["ArrHour"].isin(IN_WINDOW_HOURS)
adult["IsLWBS"] = (adult["LWBS Flag"] == "Y")
adult["IsDischarge"] = (adult["Final ED Disposition Group"] == "DISCHARGE")
adult["IsAdmit"] = (adult["Final ED Disposition Group"] == "ADMIT")

# clean any negative times (data artifacts)
time_cols = [
    "Time From Arrival To Triage","Time From Arrival To Room",
    "Time From Arrival To MD","Time From Arrival To Disposition",
    "Time From Arrival To Exit","Time From Triage To Room",
    "Time From Room To MD","Time From MD To Disposition",
    "Time From Disposition to Exit","Room to Disposition","Triage to Disposition"
]
for c in time_cols:
    adult.loc[adult[c] < 0, c] = np.nan

# ---------- Assign Split-Flow lane ----------
rng_lane = np.random.default_rng(123)

def assign_lane(row):
    if not row["InWindow"]:
        return "Off-Window"
    a = row["Acuity"]
    if a == "ESI-1":
        return "Red/Bypass"
    if a == "ESI-2":
        return "Main ED after Intake"
    if a == "ESI-3":
        # REVISED: Probabilistic blind streaming
        return "Super Track" if rng_lane.random() < 0.65 else "Main ED after Intake"
    if a in ("ESI-4","ESI-5"):
        return "Super Track"
    return "Main ED after Intake"   # ? / missing → conservative

adult["Lane"] = adult.apply(assign_lane, axis=1)

# ---------- Build simulated columns ----------
adult["sim_Arrival_To_MD"] = adult["Time From Arrival To MD"]
adult["sim_MD_To_Disp"]    = adult["Time From MD To Disposition"]
adult["sim_Disp_To_Exit"]  = adult["Time From Disposition to Exit"]
adult["sim_LWBS"]          = adult["IsLWBS"]
adult["sim_ED_LOS"]        = adult["ED LOS"]

in_win = adult["InWindow"]

# Rule 2: door-to-MD cap for in-window non-LWBS with Surge Penalty
adult["DateDate"] = adult["Arrival DateTime"].dt.date
hourly_vols = adult.groupby(["DateDate", "ArrHour"]).size()
p80_vol = hourly_vols.quantile(0.80)
rng_surge = np.random.default_rng(456)

def apply_surge_md(row):
    if not row["InWindow"] or row["IsLWBS"]:
        return row["sim_Arrival_To_MD"]
    actual = row["Time From Arrival To MD"]
    if pd.isna(actual):
        actual = DOOR_TO_MD_BASE_CAP
    
    hr_vol = hourly_vols.get((row["DateDate"], row["ArrHour"]), 0)
    if hr_vol >= p80_vol:
        sim_val = rng_surge.normal(25, 5)
        sim_val = max(10, sim_val)
    else:
        sim_val = rng_surge.normal(15, 3)
        sim_val = max(5, sim_val)
        
    return min(actual, sim_val)

adult["sim_Arrival_To_MD"] = adult.apply(apply_surge_md, axis=1)

# Rule 3: LWBS capture (deterministic retention rate => treat first N rows per lane as retained)
rng = np.random.default_rng(42)
def apply_lwbs_capture(frame, lane, keep_rate):
    idx = frame.index[(frame["Lane"] == lane) & (frame["IsLWBS"])]
    if len(idx) == 0 or keep_rate <= 0:
        return 0
    n_keep = int(round(len(idx) * keep_rate))
    keep_idx = rng.choice(idx, size=n_keep, replace=False)
    adult.loc[keep_idx, "sim_LWBS"] = False
    
    # Need to simulate their MD time using surge logic, but simplify to an avg surge/non-surge expectation
    adult.loc[keep_idx, "sim_Arrival_To_MD"] = 18 
    return n_keep

retained_st  = apply_lwbs_capture(adult, "Super Track",          LWBS_CAPTURE_SUPERTRACK)
retained_m   = apply_lwbs_capture(adult, "Main ED after Intake", LWBS_CAPTURE_MAIN_ESI2)
retained_red = apply_lwbs_capture(adult, "Red/Bypass",           LWBS_CAPTURE_RED)
print(f"  LWBS retained by Split Flow: ST={retained_st}, Main={retained_m}, Red={retained_red}")

# For retained LWBS patients, impute plausible downstream times from peers
#   arrival_to_MD already = 15.  Need MD_to_disp and disp_to_exit from peers.
def impute_retained(frame):
    for lane in ("Super Track","Main ED after Intake","Red/Bypass"):
        retained_mask = (frame["Lane"] == lane) & (frame["IsLWBS"]) & (~frame["sim_LWBS"])
        if not retained_mask.any():
            continue
        # Peers: same lane, not LWBS, have values
        peer_pool = frame[(frame["Lane"] == lane) & (~frame["IsLWBS"])]
        for acuity in frame.loc[retained_mask, "Acuity"].unique():
            sub_idx = frame.index[retained_mask & (frame["Acuity"] == acuity)]
            peer_sub = peer_pool[peer_pool["Acuity"] == acuity]
            if len(peer_sub) == 0:
                peer_sub = peer_pool
            # Retained LWBS most often would DISCHARGE from Super Track.
            if lane == "Super Track":
                peer_sub = peer_sub[peer_sub["IsDischarge"]]
            if len(peer_sub) == 0:
                peer_sub = peer_pool
            mdd_med = peer_sub["Time From MD To Disposition"].median()
            dte_med = peer_sub["Time From Disposition to Exit"].median()
            frame.loc[sub_idx, "sim_MD_To_Disp"]   = mdd_med
            frame.loc[sub_idx, "sim_Disp_To_Exit"] = dte_med
impute_retained(adult)

# Rule 4 / 5 / 6: compress MD-to-disposition (FLAT REDUCTION)
# Super Track discharged: flat 45 min reduction
m_st_d = in_win & (adult["Lane"] == "Super Track") & (adult["IsDischarge"] | (~adult["IsLWBS"] & ~adult["IsAdmit"]))
adult.loc[m_st_d, "sim_MD_To_Disp"] = np.maximum(MIN_MD_TO_DISP, adult.loc[m_st_d, "sim_MD_To_Disp"] - MD_TO_DISP_FLAT_REDUX_ST)

# Main ED (discharged or admitted): flat 20 min reduction
m_main = in_win & (adult["Lane"] == "Main ED after Intake")
adult.loc[m_main, "sim_MD_To_Disp"] = np.maximum(MIN_MD_TO_DISP, adult.loc[m_main, "sim_MD_To_Disp"] - MD_TO_DISP_FLAT_REDUX_MAIN)

# Also compress admitted Super Track
m_st_adm = in_win & (adult["Lane"] == "Super Track") & adult["IsAdmit"]
adult.loc[m_st_adm, "sim_MD_To_Disp"] = np.maximum(MIN_MD_TO_DISP, adult.loc[m_st_adm, "sim_MD_To_Disp"] - MD_TO_DISP_FLAT_REDUX_MAIN)

# Recompute sim_ED_LOS:
#   Preferred: sim_arrival_to_MD + sim_MD_to_disp + sim_disp_to_exit (when all present)
#   Fallback:  actual ED LOS - reduction in arrival_to_MD - reduction in MD_to_disp
#   Clamp savings to actual LOS (cannot save more time than the patient spent in the ED)
def safe_sum(r):
    vals = [r["sim_Arrival_To_MD"], r["sim_MD_To_Disp"], r["sim_Disp_To_Exit"]]
    if all(pd.notna(v) for v in vals):
        return sum(vals)
    return np.nan

sim_computed = adult.apply(safe_sum, axis=1)

# Savings in each piece (can't be negative — simulation never lengthens the visit)
sav_md  = (adult["Time From Arrival To MD"]   - adult["sim_Arrival_To_MD"]).clip(lower=0).fillna(0)
sav_mdd = (adult["Time From MD To Disposition"] - adult["sim_MD_To_Disp"]).clip(lower=0).fillna(0)
total_sav = sav_md + sav_mdd

# Cap total savings so sim LOS doesn't drop below 30 min (physical floor for any ED visit)
MIN_SIM_LOS = 30
headroom = (adult["ED LOS"] - MIN_SIM_LOS).clip(lower=0)
total_sav_capped = np.minimum(total_sav, headroom)

fallback = adult["ED LOS"] - total_sav_capped

# Use computed sim when valid AND not crazy low; else fallback
use_computed = sim_computed.notna() & (sim_computed >= MIN_SIM_LOS) & (sim_computed <= adult["ED LOS"])
adult["sim_ED_LOS"] = np.where(use_computed, sim_computed, fallback)

# Off-window: no change
adult.loc[~in_win, "sim_ED_LOS"]           = adult.loc[~in_win, "ED LOS"]
adult.loc[~in_win, "sim_Arrival_To_MD"]    = adult.loc[~in_win, "Time From Arrival To MD"]
adult.loc[~in_win, "sim_MD_To_Disp"]       = adult.loc[~in_win, "Time From MD To Disposition"]

# Final guard: sim LOS never negative, never > actual
adult["sim_ED_LOS"] = adult["sim_ED_LOS"].clip(lower=0)
adult["sim_ED_LOS"] = np.minimum(adult["sim_ED_LOS"], adult["ED LOS"])

# ---------- Summaries ----------
def pct(n, d): return 0 if d == 0 else 100 * n / d

def stats(series):
    s = series.dropna()
    if len(s) == 0:
        return dict(n=0, median=np.nan, p90=np.nan, mean=np.nan)
    return dict(n=len(s), median=s.median(), p90=s.quantile(0.90), mean=s.mean())

# Summary 1: Headline KPIs
kpi_rows = []
def kpi(name, actual, simulated, unit):
    kpi_rows.append([name, actual, simulated, simulated - actual if isinstance(actual,(int,float)) else "", unit])

# Door-to-Clinician
a_md  = adult["Time From Arrival To MD"]
s_md  = adult["sim_Arrival_To_MD"]
kpi("Door-to-Clinician — Median", a_md.median(), s_md.median(), "min")
kpi("Door-to-Clinician — P90",    a_md.quantile(0.9), s_md.quantile(0.9), "min")

# LWBS
a_lwbs = adult["IsLWBS"].sum()
s_lwbs = adult["sim_LWBS"].sum()
kpi("LWBS — Count",       a_lwbs, s_lwbs, "encounters")
kpi("LWBS — Rate",        pct(a_lwbs, len(adult)), pct(s_lwbs, len(adult)), "%")

# Discharged LOS (exclude LWBS)
a_disc_mask = adult["IsDischarge"]
s_disc_mask = adult["IsDischarge"] | (adult["IsLWBS"] & ~adult["sim_LWBS"])  # retained LWBS mostly discharge from ST
a_disch_los = adult.loc[a_disc_mask, "ED LOS"]
s_disch_los = adult.loc[s_disc_mask, "sim_ED_LOS"]
kpi("Discharged LOS — Median (hr)", a_disch_los.median()/60, s_disch_los.median()/60, "hr")
kpi("Discharged LOS — P90 (hr)",    a_disch_los.quantile(0.9)/60, s_disch_los.quantile(0.9)/60, "hr")

# Admitted LOS
a_adm = adult.loc[adult["IsAdmit"], "ED LOS"]
s_adm = adult.loc[adult["IsAdmit"], "sim_ED_LOS"]
kpi("Admitted LOS — Median (hr)", a_adm.median()/60, s_adm.median()/60, "hr")
kpi("Admitted LOS — P90 (hr)",    a_adm.quantile(0.9)/60, s_adm.quantile(0.9)/60, "hr")

kpi("Total ADULT ED Encounters", len(adult), len(adult), "")
kpi("In-Window Encounters (08:00–01:00)", int(adult["InWindow"].sum()), int(adult["InWindow"].sum()), "")
kpi("LWBS Retained by Split Flow", 0, int(a_lwbs - s_lwbs), "encounters")

kpi_df = pd.DataFrame(kpi_rows, columns=["Metric","Actual","Simulated (Split Flow)","Delta","Unit"])

# Summary 2: Volume by Lane
lane_df = adult.groupby("Lane").agg(
    encounters=("Encounter # (CSN)", "count"),
    avg_daily=("Encounter # (CSN)", lambda x: len(x)/293),
    disch_pct=("IsDischarge", lambda x: 100*x.mean()),
    admit_pct=("IsAdmit",     lambda x: 100*x.mean()),
    lwbs_pct_actual=("IsLWBS", lambda x: 100*x.mean()),
    arrival_to_md_med_actual=("Time From Arrival To MD", "median"),
    ed_los_med_actual=("ED LOS", "median"),
).round(2).reset_index()

sim_by_lane = adult.groupby("Lane").agg(
    arrival_to_md_med_sim=("sim_Arrival_To_MD","median"),
    lwbs_pct_sim=("sim_LWBS", lambda x:100*x.mean()),
    ed_los_med_sim=("sim_ED_LOS","median"),
).round(2).reset_index()
lane_df = lane_df.merge(sim_by_lane, on="Lane")

# Summary 3: LWBS by Acuity
lwbs_acuity = adult.groupby("Acuity").agg(
    encounters=("Encounter # (CSN)","count"),
    lwbs_actual=("IsLWBS", "sum"),
    lwbs_sim=("sim_LWBS", "sum"),
).reset_index()
lwbs_acuity["lwbs_rate_actual_%"] = (100*lwbs_acuity["lwbs_actual"]/lwbs_acuity["encounters"]).round(2)
lwbs_acuity["lwbs_rate_sim_%"]    = (100*lwbs_acuity["lwbs_sim"]/lwbs_acuity["encounters"]).round(2)
lwbs_acuity["retained_by_SF"]     = lwbs_acuity["lwbs_actual"] - lwbs_acuity["lwbs_sim"]

# Summary 4: Targets vs Simulated Performance
targets = pd.DataFrame([
    ["Door-to-Clinician — Median","≤15 min initial / ≤10 min steady", s_md.median()],
    ["Door-to-Clinician — P90","≤45 min initial / ≤30 min steady",    s_md.quantile(0.9)],
    ["LWBS Rate","≤3.0% initial / ≤1.0% steady",                       pct(s_lwbs, len(adult))],
    ["Discharged LOS — Median","≤4.5 hr initial / ≤4.0 hr steady",    s_disch_los.median()/60],
    ["Discharged LOS — P90","≤7.0 hr initial / ≤6.5 hr steady",        s_disch_los.quantile(0.9)/60],
], columns=["Metric","Split-Flow Target","Simulated Value"])

def status(metric, val):
    if "Door" in metric and "Median" in metric: return "Meets initial" if val <= 15 else "Below target"
    if "Door" in metric and "P90"    in metric: return "Meets initial" if val <= 45 else "Below target"
    if "LWBS"  in metric:                        return "Meets initial" if val <= 3.0 else "Below target"
    if "Disch" in metric and "Median" in metric: return "Meets initial" if val <= 4.5 else "Below target"
    if "Disch" in metric and "P90"    in metric: return "Meets initial" if val <= 7.0 else "Below target"
    return ""

targets["Status"] = targets.apply(lambda r: status(r["Metric"], r["Simulated Value"]), axis=1)
targets["Simulated Value"] = targets["Simulated Value"].round(2)

# Summary 5: Safety check — ESI-3 sent to Super Track that were actually admitted
# (by our lane rules this is 0; but add check anyway for any ESI-3 ST whose actual outcome admit)
st_pool = adult[(adult["Lane"] == "Super Track")]
safety = pd.DataFrame([
    ["Super Track encounters (simulated)", len(st_pool)],
    ["  of which actual = ADMIT (pull-back cases)", int(st_pool["IsAdmit"].sum())],
    ["  of which actual = DISCHARGE",               int(st_pool["IsDischarge"].sum())],
    ["  of which actual = LWBS",                    int(st_pool["IsLWBS"].sum())],
    ["Super Track pull-back rate (% admits in ST)", f"{100*st_pool['IsAdmit'].mean():.2f}%" ],
], columns=["Safety Check","Value"])

# Summary 6: Sensitivity analysis
sens_rows = []
for cap in [10, 15, 20, 30]:
    for keep in [0.60, 0.75, 0.85, 0.95]:
        # quick recompute
        tmp_md = adult["Time From Arrival To MD"].copy()
        tmp_md.loc[in_win & (~adult["IsLWBS"])] = np.minimum(
            tmp_md.loc[in_win & (~adult["IsLWBS"])].fillna(cap), cap)
        # LWBS
        st_lwbs_idx = adult.index[(adult["Lane"]=="Super Track") & (adult["IsLWBS"])]
        main_lwbs_idx = adult.index[(adult["Lane"]=="Main ED after Intake") & (adult["IsLWBS"])]
        red_lwbs_idx  = adult.index[(adult["Lane"]=="Red/Bypass") & (adult["IsLWBS"])]
        retained = int(round(len(st_lwbs_idx)*keep)) + int(round(len(main_lwbs_idx)*LWBS_CAPTURE_MAIN_ESI2)) + int(round(len(red_lwbs_idx)*LWBS_CAPTURE_RED))
        sim_lwbs_rate = pct(a_lwbs - retained, len(adult))
        sens_rows.append([cap, keep, tmp_md.median(), tmp_md.quantile(0.9), sim_lwbs_rate])
sens_df = pd.DataFrame(sens_rows, columns=[
    "Door-to-MD Cap (min)","Super Track LWBS Capture",
    "Sim Door-to-MD Median","Sim Door-to-MD P90","Sim LWBS Rate %"])

# Summary 7: Executive summary text
exec_rows = [
    ["Study period","2025-07-01 to 2026-04-19 (293 days)"],
    ["Adult ED encounters","{:,}".format(len(adult))],
    ["Average daily volume","{:.1f}".format(len(adult)/293)],
    ["In-window volume (08:00–01:00)","{:,} ({:.1f}%)".format(int(adult['InWindow'].sum()), 100*adult['InWindow'].mean())],
    ["",""],
    ["ACTUAL — Door-to-Clinician (median / P90)","{:.0f} / {:.0f} min".format(a_md.median(), a_md.quantile(0.9))],
    ["SIMULATED — Door-to-Clinician (median / P90)","{:.0f} / {:.0f} min".format(s_md.median(), s_md.quantile(0.9))],
    ["",""],
    ["ACTUAL LWBS rate","{:.2f}% ({} encounters)".format(100*adult['IsLWBS'].mean(), int(adult['IsLWBS'].sum()))],
    ["SIMULATED LWBS rate","{:.2f}% ({} encounters)".format(pct(s_lwbs,len(adult)), int(s_lwbs))],
    ["LWBS reduction under Split Flow","{} encounters ({:.1f}% relative)".format(int(a_lwbs-s_lwbs), 100*(a_lwbs-s_lwbs)/max(a_lwbs,1))],
    ["",""],
    ["ACTUAL Discharged LOS (median / P90)","{:.2f} / {:.2f} hr".format(a_disch_los.median()/60, a_disch_los.quantile(0.9)/60)],
    ["SIMULATED Discharged LOS (median / P90)","{:.2f} / {:.2f} hr".format(s_disch_los.median()/60, s_disch_los.quantile(0.9)/60)],
    ["",""],
    ["ACTUAL Admitted LOS (median / P90)","{:.2f} / {:.2f} hr".format(a_adm.median()/60, a_adm.quantile(0.9)/60)],
    ["SIMULATED Admitted LOS (median / P90)","{:.2f} / {:.2f} hr".format(s_adm.median()/60, s_adm.quantile(0.9)/60)],
    ["Note","Admit LOS is dominated by downstream boarding; Split Flow does not change it."],
]
exec_df = pd.DataFrame(exec_rows, columns=["Item","Value"])

# ---------- Write Excel ----------
print(f"Writing results to: {OUT}")
encounter_out = adult[[
    "Encounter # (CSN)","Arrival DateTime","ArrHour","InWindow","Acuity","Lane",
    "Final ED Disposition Group","IsLWBS","sim_LWBS",
    "Time From Arrival To MD","sim_Arrival_To_MD",
    "Time From MD To Disposition","sim_MD_To_Disp",
    "Time From Disposition to Exit","sim_Disp_To_Exit",
    "ED LOS","sim_ED_LOS",
]].rename(columns={
    "Time From Arrival To MD":"Actual_Arrival_To_MD_min",
    "sim_Arrival_To_MD":"Sim_Arrival_To_MD_min",
    "Time From MD To Disposition":"Actual_MD_To_Disp_min",
    "sim_MD_To_Disp":"Sim_MD_To_Disp_min",
    "Time From Disposition to Exit":"Actual_Disp_To_Exit_min",
    "sim_Disp_To_Exit":"Sim_Disp_To_Exit_min",
    "ED LOS":"Actual_ED_LOS_min",
    "sim_ED_LOS":"Sim_ED_LOS_min",
})

with pd.ExcelWriter(OUT, engine="openpyxl") as w:
    exec_df.to_excel(w, sheet_name="1. Executive Summary", index=False)
    kpi_df.to_excel(w,  sheet_name="2. Headline KPIs", index=False)
    targets.to_excel(w, sheet_name="3. vs Split-Flow Targets", index=False)
    lane_df.to_excel(w, sheet_name="4. By Lane", index=False)
    lwbs_acuity.to_excel(w, sheet_name="5. LWBS by Acuity", index=False)
    safety.to_excel(w,  sheet_name="6. Safety Check", index=False)
    sens_df.to_excel(w, sheet_name="7. Sensitivity", index=False)
    encounter_out.to_excel(w, sheet_name="8. Encounter-Level", index=False)

# Format: column widths
from openpyxl import load_workbook
wb = load_workbook(OUT)
for sh in wb.sheetnames:
    ws = wb[sh]
    for col_cells in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        col_letter = col_cells[0].column_letter
        ws.column_dimensions[col_letter].width = min(max(length + 2, 10), 55)
wb.save(OUT)

# ---------- Console summary ----------
print()
print("="*78)
print("SPLIT-FLOW BACK TEST — HEADLINE RESULTS")
print("="*78)
print(kpi_df.to_string(index=False))
print()
print("TARGETS vs SIMULATED")
print(targets.to_string(index=False))
print()
print("BY LANE")
print(lane_df.to_string(index=False))
print()
print("LWBS BY ACUITY")
print(lwbs_acuity.to_string(index=False))
print()
print(f"Full results workbook: {OUT.name}")
